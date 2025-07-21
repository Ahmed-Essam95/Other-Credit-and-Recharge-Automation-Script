from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl as excel
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
import time
import os


accounts_sheet_path = r" File Path .xlsx"
accounts_workbook = excel.load_workbook(accounts_sheet_path)
source_sheet = accounts_workbook["Sheet1"]
save_sheet = accounts_workbook["CX"]


# Without GUI
# from selenium.webdriver.chrome.options import Options
# options_list = Options()
# options_list.add_argument("--headless")


occ_robot = webdriver.Chrome()
occ_robot.maximize_window()
occ_robot.get(" Path of the web app ")
hold = WebDriverWait( occ_robot ,20)

level = 0
account_performance = ""
un_bill = ""

def login_page(username,password) :
    """Login FX"""
    # Check Point To Move.
    hold.until(EC.visibility_of_element_located((By.XPATH,"//div[@id='footer']//p")))

    # Enter UserName
    hold.until(EC.visibility_of_element_located((By.ID,"username"))).send_keys(username)
    # Enter Password
    hold.until(EC.visibility_of_element_located((By.ID, "password"))).send_keys(password)
    # Press Ok To log in
    hold.until(EC.element_to_be_clickable((By.CLASS_NAME, "Button Standard"))).click()
    # Press Customers Button
    hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Customers"))).click()




def occ_cycle(acc_num):
    global account_performance,un_bill
    account_occ = []
    try:
        def page_source(acc_num):
            """Enter account number to pass source page"""

            # Press Search to view source fields
            hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Search"))).click()

            # Check New Page Elements visibility
            hold.until(EC.visibility_of_element_located((By.XPATH, "//div[@id='footer']//p")))

            hold.until(EC.visibility_of_element_located((By.ID, "ID"))).send_keys(acc_num)


            # Press Search.
            hold.until(EC.element_to_be_clickable((By.ID, " Search_Button"))).click()
            time.sleep(0.75)
            # Pres to enter the account
            hold.until(EC.element_to_be_clickable((By.CLASS_NAME, "DATblTDALinkTxt"))).click()

        page_source(acc_num)

        def re_fetch():
            """Re-Fetching Home Page For Each Call"""
            # Check Point To Move / Validation For All Account Dials Before Go
            return hold.until(
                EC.visibility_of_all_elements_located((By.XPATH, "//table[@id='TableModel']//tbody//tr")))

        re_fetch()

        # Get Unbilled Data
        try:
            element_name = hold.until(EC.visibility_of_element_located(
                (By.XPATH, "//*[@id='ID Name']/table[1]/tbody/tr[3]/td[3]/div/span"))).text
            if element_name == "Unbilled usage" :
                un_billed_usage = hold.until(EC.visibility_of_element_located((By.XPATH,"//*[@id='ID Name']/table[1]/tbody/tr[3]/td[4]/div/span"))).text
                un_bill = un_billed_usage

        except:
            un_bill = "Null Data"




        time.sleep(0.25)
        hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Billing and rating"))).click()
        time.sleep(0.10)
        hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Other credits and charges"))).click()


        try :
            # Check Point To Move.
            hold.until(EC.visibility_of_element_located((By.XPATH, "//div[@id='footer']//p")))
            time.sleep(0.5)
            occ_kount = WebDriverWait( occ_robot ,1).until(EC.visibility_of_all_elements_located((By.XPATH, "//*[@id='OCC Table Model']/tbody/tr")))

            for row_occ in occ_kount:
                if row_occ.find_element(By.XPATH, ".//td[3]").text == "Monthly fees adjustment" :
                    one_occ = []
                    amount = row_occ.find_element(By.XPATH, ".//td[contains(text() , 'EGP')]").text
                    if "(" not in amount and ")" not in amount :
                        continue
                    else:
                        pure_amount = ""
                        for ch in amount:
                            if ch in ["(", ")", " ", "E", "G", "P"]:
                                continue
                            else:
                                pure_amount += ch
                        times = row_occ.find_element(By.XPATH, ".//td[6]").text
                        service = row_occ.find_element(By.XPATH, ".//td[3]").text
                        comment = row_occ.find_element(By.XPATH, ".//td[8]").text

                        account_num_occ = hold.until(EC.visibility_of_element_located((By.XPATH,"//span[contains(text() , 'Customer code')]//following-sibling::span"))).text


                        if not comment :
                            comment = "Null Data"
                        valid_from = row_occ.find_element(By.XPATH, ".//td[9]").text
                        one_occ.append(acc_num)
                        one_occ.append(pure_amount)
                        one_occ.append(times)
                        one_occ.append(service)
                        one_occ.append(comment)
                        one_occ.append(valid_from)
                        one_occ.append(len(occ_kount))
                        one_occ.append(account_num_occ)

                        one_occ.append(un_bill)


                        account_occ.append(one_occ)

                else:
                    continue

            time.sleep(0.5)
            hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Search"))).click()
            time.sleep(0.5)

        except :
            account_occ.append([acc_num, "N/A", "N/A", "N/A", "N/A", "N/A", hold.until(EC.visibility_of_element_located((By.XPATH,"//span[contains(text() , 'Customer code')]//following-sibling::span"))).text,"N/A",un_bill])
            hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Search"))).click()

    except Exception as e:
        account_occ.append(["Not Done","Not Done","Not Done","Not Done","Not Done","Not Done","Not Done","Not Done"])
        account_performance = "Not Done"
        occ_robot.save_screenshot(f"PPT FLAG{acc_num}.png")

        print(f"Issued Account # {acc_num}")
        print(f"Error : {e}")

        point_zero = hold.until(EC.element_to_be_clickable((By.LINK_TEXT, "Search")))
        point_zero.click()


    else:
        account_performance = "Done"


    return account_occ



# Start Engine.
login_page(" User Name "," Password")


for account_cl in range( 2,  source_sheet.max_row+1 ) :
    level += 1

    account_number = str(source_sheet.cell(account_cl,1).value)

    print(f"Level : {level}, Account Number : {account_number}")

    result_account_occ = occ_cycle(account_number)

    source_sheet.cell(account_cl,2).value = account_performance




    for each_row in result_account_occ :
        save_sheet.append(each_row)
    accounts_workbook.save(accounts_sheet_path)




print("Saving Data....")
occ_robot.quit()
print("Task has been performed.")


